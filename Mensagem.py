import pandas as pd
import openpyxl
import os
from openpyxl import load_workbook
from openpyxl import workbook
import webbrowser as web
import keyboard 
import pyautogui
import time

#df  = pd.read_excel("C:\\Meus Projetos\\projetosFaculdade\\Planilhateste.xlsx")


df = pd.read_excel("C:\\TESTE GIT 2\\projetosFaculdade\\Planilhateste.xlsx", index_col=0)
df.head(100)

planilha = load_workbook(filename= 'Planilhateste.xlsx')
plan = planilha.active




class Mensagem:

    def mensagem():

        print(df)
        print("\n\n" + "Digite a coluna onde o telefone está:" + "\n")
        coluna = int(input())
        print("\n" + "Digite a linha onde o telefone está:" + "\n")
        linha = int(input())
        print("\n" + "Digite o nome da pessoa da qual a mensagem será enviada: " + "\n")
        nome = input()
        Varmeng = plan.cell(row=linha,column=coluna).value 
        print("\n" + "Dado da Célula: " + Varmeng + "\n\n" + "Nome digitado: " + nome)
        print("\n" + "Os dados estão corretos?: " + "\n")
        corr = input().upper()
        if corr== "SIM":
            url = 'https://web.whatsapp.com/send?phone='
            #tel = Varmeng
            tel = '+55 11961611974' 
            msg = "Olá, " + "\n" + nome + ", tudo bem?" + "\n" + " Somos da Digital, da *CAIXA*." + "\n" + "Você pode RENOVAR o seu empréstimo *CONSIGNADO* CAIXA, com taxas e prazos SUPER ESPECIAIS! A parcela será mantida, e você ainda receberá um troco na RENOVAÇÃO." + "\n" + " Gostou?" + "\n" + " retorne essa mensagem."
            print("\n" + "Pressione Enter para enviar a mensagem:" + "\n")
            keyboard.wait("enter")
            ##web.open(url+tel,msg)
            web.open('https://web.whatsapp.com/send?phone='+tel+'&text='+msg)
            time.sleep(13)
            pyautogui.hotkey('enter', presses = 1)

      