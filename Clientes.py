import pandas as pd
import openpyxl
import os
from openpyxl import load_workbook

df  = pd.read_excel("C:\\Meus Projetos\\projetosFaculdade\\Planilhateste.xlsx")
#df = pd.read_excel("C:\\TESTE GIT 2\\projetosFaculdade\\Planilhateste.xlsx")
df.head(41)

plan = load_workbook(filename= 'Planilhateste.xlsx')



class VisualisarCad:

    @staticmethod
    def clientes():
        return print(df.head(41))
    
class Escrever:

    def alterar():
        return         #os.system("cls")
        print("Escolha a coluna a ser alterada:")
        coluna = input()
        print("\n" + "Escolha a linha a ser alterada:")
        linha  = int(input())
        print("\n Escolha o novo valor a ser inserido:")
        valor  = input()
        plan.cell(row=linha,colunm=coluna,value=valor)
        plan.template = True
        plan.save(plan)
    
