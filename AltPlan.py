import pandas as pd
import openpyxl
import os
from openpyxl import load_workbook
from openpyxl import workbook


df  = pd.read_excel("C:\\Meus Projetos\\projetosFaculdade\\Planilhateste.xlsx")
#df = pd.read_excel("C:\\TESTE GIT 2\\projetosFaculdade\\Planilhateste.xlsx")
df.head(41)

planilha = load_workbook(filename= 'Planilhateste.xlsx')
plan = planilha.active

class Escrever:

    def alterar():
        os.system("cls")
        print("Escolha a coluna a ser alterada:" + "\n")
        coluna = int(input())
        print("\n" + "Escolha a linha a ser alterada:" + "\n")
        linha  = int(input())
        print("\n" + "Escolha o novo valor a ser inserido:" + "\n")
        valor  = input()
        os.system("cls")
        print("---------------------------------------Valores Atualizados------------------------------------------" + "\n")

        plan.cell(row=linha,column=coluna,value=valor)
        plan.template = True
        planilha.save('Planilhateste.xlsx')
        