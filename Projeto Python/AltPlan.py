import pandas as pd
import os
import time
from openpyxl import load_workbook

df = pd.read_excel("C:\\planilha.xlsx", index_col=0)
#df = pd.read_excel("C:\\TESTE GIT 2\\projetosFaculdade\\Projeto Python\\Planilhateste.xlsx", index_col=0)
#df = pd.read_excel("C:\\Users\\leonardo.gomes\\Desktop\Meus Projetos\\projetosFaculdade\\Projeto Python\\Planilhateste.xlsx",index_col=0)

planilha = load_workbook(filename="C:\\planilha.xlsx")
plan = planilha.active

class Escrever:

    def alterar():
      
      while True:
        os.system("cls")
        while True:
           try:  
             coluna = int(input("Escolha a coluna a ser alterada (somente números!!):" + "\n"))
             break
           except ValueError:
              print('\n' + "Digite apenas números!!")
              time.sleep(2)
              os.system("cls")
        while True:
           try:      
             linha = int(input("\n" + "Escolha a linha a ser alterada (somente números!!):" + "\n"))
             break
           except ValueError:
             print("\n" + "Digite apenas números!!" + "\n")
             time.sleep(2)
             os.system("cls")
        
        print("\n" + "Escolha o novo valor a ser inserido:" + "\n")
        valor = input()

        os.system('cls')
        print('Os dados inseridos estão corretos? (Sim/Nao): Coluna ', coluna, 'Linha ', linha, 'Valor ', valor + '\n')
        escolha = input().upper()

        if escolha == 'SIM':
            print("---------------------------------------Valores Atualizados------------------------------------------" + "\n")
            time.sleep(1.5)
            plan.cell(row=linha, column=coluna, value=valor)
            planilha.save('planilha.xlsx')
            break
        elif escolha == 'NAO' or escolha == 'NÃO':
            os.system('cls')
            print("---------------------------------------Aguarde------------------------------------------" + "\n")
            time.sleep(3.0)
        else:
            os.system('cls')
            print("---------------------------------------Opção inválida. Digite *Sim* ou *Nao*------------------------------------------" + "\n")
            time.sleep(3.0)